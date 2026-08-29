#!/usr/bin/env python3
"""Genera prensa.html desde Antecedente_Prensa_ConsolidadoMDD.xlsx.

Uso: python3 tools/gen_prensa.py <Prensa.xlsx>
Filas con URL real quedan enlazadas; el resto se lista con medio y fecha.
"""
import sys, html
from openpyxl import load_workbook

HEAD = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Prensa · Matías Díaz</title>
<meta name="description" content="Apariciones en prensa de Matías Díaz (DIE USACH): electromovilidad, carga de vehículos eléctricos y almacenamiento, 2014-2026.">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Instrument+Serif:ital@0;1&family=Archivo:wght@400;500;600;700&display=swap">
<link rel="stylesheet" href="assets/site.css">
</head>
<body>
<nav><div class="wrap">
  <a class="logo" href="index.html">Matías <b>Díaz</b></a>
  <a href="investigacion.html">Investigación</a>
  <a href="publicaciones.html">Publicaciones</a>
  <a href="videos.html">Videos</a>
  <a class="on" href="prensa.html">Prensa</a>
  <a href="docencia.html">Docencia</a>
  <a href="equipo.html">Equipo</a>
  <a href="postula.html">Postula</a>
  <a href="cv.html">CV</a>
</div></nav>
<header class="pagehead wrap">
  <h1>Prensa</h1>
  <p>{total} apariciones registradas entre 2014 y 2026 en prensa escrita, radio y televisión, sobre electromovilidad, carga de vehículos eléctricos, baterías y política energética.</p>
</header>
<section style="padding-top:40px"><div class="wrap">
  <div class="head"><h2>Destacadas</h2></div>
  <div class="quotes">
    <a class="q" href="https://www.cooperativaciencia.cl/tecnologia/2026/03/13/baterias-reutilizadas-de-autos-electricos-energizaran-espacio-de-lollapalooza-2026/" target="_blank" rel="noopener"><p>Baterías reutilizadas de autos eléctricos energizarán Lollapalooza 2026</p><span>Cooperativa Ciencia · 2026</span></a>
    <div class="q"><p>Ya es ley la normativa para convertir autos a combustión en eléctricos</p><span>CNN Chile · 2025</span></div>
    <a class="q" href="https://www.biobiochile.cl/especial/aqui-tierra/noticias/2024/10/29/ingenieros-del-e2tech-de-la-usach-desarrollan-el-primer-cargador-de-autos-electricos-hecho-en-chile.shtml" target="_blank" rel="noopener"><p>El primer cargador de autos eléctricos hecho en Chile</p><span>BioBioChile · 2024</span></a>
    <div class="q"><p>Entrevista sobre el pago de permisos de circulación de autos eléctricos</p><span>24 Horas TVN · 2025</span></div>
    <a class="q" href="https://www.usach.cl/news/usach-se-adjudica-proyecto-corfo-4-millones-dolares" target="_blank" rel="noopener"><p>USACH se adjudica proyecto CORFO de 4 millones de dólares</p><span>USACH · 2024</span></a>
    <div class="q"><p>El potencial del retrofit en Chile, columna de opinión</p><span>Cooperativa.cl · 2025</span></div>
  </div>
</div></section>
<section class="alt"><div class="wrap">
  <div class="head"><h2>Registro completo</h2><p>Índice consolidado presentado en la postulación a Profesor Titular (junio de 2026).</p></div>
"""

FOOT = """</div></section>
<footer><div class="wrap">
  <span>© 2026 Matías Díaz · DIE USACH</span>
  <span>Generado desde el índice consolidado de prensa</span>
</div></footer>
<script src="assets/site.js"></script>
</body>
</html>
"""

def esc(x):
    return html.escape(str(x).strip()) if x else ""

def main(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Apariciones en Prensa"]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[2] and r[5]]
    out = [HEAD.format(total=len(rows))]
    years = sorted({int(r[2]) for r in rows}, reverse=True)
    for y in years:
        out.append(f'<div class="pubyear">{y}</div><div class="scroll"><table>')
        out.append("<tr><th>Fecha</th><th>Medio</th><th>Título</th><th>Tema</th></tr>")
        for r in rows:
            if int(r[2]) != y:
                continue
            _, fecha, _, medio, tipo, titulo, tema, alcance, url = (list(r) + [None]*9)[:9]
            t = esc(titulo)
            if url and str(url).startswith("http"):
                t = f'<a href="{esc(url)}" target="_blank" rel="noopener">{t}</a>'
            out.append(f"<tr><td>{esc(fecha)}</td><td>{esc(medio)}</td><td>{t}</td><td>{esc(tema)}</td></tr>")
        out.append("</table></div>")
    out.append(FOOT)
    with open("prensa.html", "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print(f"OK: {len(rows)} apariciones, {len(years)} años")

if __name__ == "__main__":
    main(sys.argv[1])
