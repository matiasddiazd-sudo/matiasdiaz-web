#!/usr/bin/env python3
"""Genera publicaciones.html y publications.bib desde Publicaciones_ListadoIEEE_v4.xlsx.

La planilla v4 (agosto 2026, verificada por el autor) trae 135 referencias en formato
IEEE con año, tipo (Revista/Conferencia) y DOI. Este script las vuelca a la página
con el nombre propio en negrita, la revista o conferencia en cursiva y enlace DOI.

Uso: python3 tools/gen_publicaciones_v4.py <Publicaciones_ListadoIEEE_v4.xlsx>
"""
import sys, html, re, json
from openpyxl import load_workbook

NAME_RE = re.compile(r"(M\.\s?D[ií]az(?:\sD[ií]az)?|Mat[ií]as\sD[ií]az(?:\sD[ií]az)?|D[ií]az,?\s?M\.?|M\sD[ií]az)")

HEAD = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Publicaciones · Matías Díaz</title>
<meta name="description" content="Publicaciones de Matías Díaz (DIE USACH): artículos en revistas WoS-JCR y conferencias internacionales, con enlace DOI.">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Instrument+Serif:ital@0;1&family=Archivo:wght@400;500;600;700&display=swap">
<link rel="stylesheet" href="assets/site.css">
</head>
<body>
<nav><div class="wrap">
  <a class="logo" href="index.html">Matías <b>Díaz</b></a>
  <a href="investigacion.html">Investigación</a>
  <a class="on" href="publicaciones.html">Publicaciones</a>
  <a href="prensa.html">Prensa y videos</a>
  <a href="docencia.html">Docencia</a>
  <a href="equipo.html">Equipo</a>
  <a href="cv.html">CV</a>
</div></nav>
<header class="pagehead wrap">
  <p>{n_t} publicaciones científicas: {n_j} artículos en revistas y {n_c} en conferencias internacionales, 2013-2026. Listado verificado por el autor (agosto de 2026). Perfil completo en <a href="https://scholar.google.com/citations?user=-43YaJIAAAAJ" target="_blank" rel="noopener">Google Scholar</a> · <a href="publications.bib">BibTeX</a>.</p>
</header>
"""
HEAD = HEAD.replace('<header class="pagehead wrap">', '<header class="pagehead wrap">\n  <h1>Publicaciones</h1>')

FOOT = """
<footer><div class="wrap">
  <span>© 2026 Matías Díaz · DIE USACH</span>
  <span>Listado verificado por el autor · agosto de 2026</span>
</div></footer>
<script src="assets/site.js"></script>
</body>
</html>
"""

def doi_url(doi):
    if not doi:
        return None
    doi = doi.strip()
    if doi.startswith('http'):
        return doi
    if doi.lower().startswith('doi.org'):
        return 'https://' + doi
    if doi.startswith('10.'):
        return 'https://doi.org/' + doi
    return None

def venue_split(tail):
    """tail = texto tras el cierre de comillas del título. Devuelve (venue, resto)."""
    cut = len(tail)
    m = re.search(r',\s*(vol\.|no\.|pp\.|19\d\d|20\d\d)', tail)
    if m:
        cut = min(cut, m.start())
    p = tail.find('),')
    if p != -1:
        cut = min(cut, p + 1)
    return tail[:cut], tail[cut:]

def render_ref(ref):
    e = html.escape(ref, quote=False)
    # separa autores “título,” cola
    m = re.match(r'^(.*?)[“"](.+?)[”"],?\s*(.*)$', e, re.S)
    if m:
        auth, title, tail = m.groups()
        title = title.rstrip(', ')
        venue, rest = venue_split(tail)
        out = f'{auth}“{title},” <em>{venue}</em>{rest}'
    else:
        out = e
    out = NAME_RE.sub(lambda x: f'<b>{x.group(1)}</b>', out)
    return out

def thumb_for(ref, tipo):
    m = re.search(r'[”"],?\s*(.*)$', ref, re.S)
    tail = m.group(1) if m else ''
    if tipo == 'Conferencia':
        p = re.search(r'\(([A-Z][A-Za-z\- ]{2,18})\)', tail)
        if p:
            return p.group(1).split()[0]
        for acr in ('EVER', 'CHILECON', 'CONCAPAN', 'ICA-ACCA', 'SMART', 'IECON', 'ISIE',
                    'ICIT', 'SPEC', 'ECCE', 'APEC', 'COBEP', 'ARGENCON', 'INGELECTRA', 'PEDG'):
            if acr in tail:
                return acr
        return 'Conf.'
    for key, lab in [('IEEE', 'IEEE'), ('Processes', 'Processes'), ('Energies', 'Energies'),
                     ('Machines', 'Machines'), ('Electronics', 'MDPI'), ('IET', 'IET'),
                     ('Journal', 'Journal'), ('Results in', 'Elsevier')]:
        if key in tail:
            return lab
    return 'Revista'

def bib_key(ref, anio, used):
    m = re.search(r'[“"](.+?)[”"]', ref)
    words = re.sub(r'[^a-z ]', '', (m.group(1) if m else 'pub').lower()).split()
    stop = {'a', 'an', 'the', 'of', 'for', 'and', 'in', 'on', 'with', 'to', 'based', 'using'}
    sig = [w for w in words if w not in stop][:3]
    key = f"diaz{anio}" + ''.join(w[:4] for w in sig)
    while key in used:
        key += 'x'
    used.add(key)
    return key

def bib_entry(r, used):
    ref = r['ref']
    m = re.match(r'^(.*?)[“"](.+?)[”"],?\s*(.*)$', ref, re.S)
    auth = title = tail = ''
    if m:
        auth, title, tail = m.groups()
        title = title.rstrip(', ')
        auth = auth.rstrip(', ').strip()
    venue, rest = venue_split(tail) if m else ('', '')
    kind = 'article' if r['tipo'] == 'Revista' else 'inproceedings'
    vfield = 'journal' if kind == 'article' else 'booktitle'
    key = bib_key(ref, r['anio'], used)
    lines = [f"@{kind}{{{key},"]
    if auth:
        lines.append(f"  author = {{{auth.replace(' and ', ' AND ').replace(', ', ' AND ').replace(' AND AND ', ' AND ')}}},")
    lines.append(f"  title = {{{title}}},")
    if venue.strip(' ,'):
        lines.append(f"  {vfield} = {{{venue.strip(' ,')}}},")
    lines.append(f"  year = {{{r['anio']}}},")
    u = doi_url(r['doi'])
    if r['doi'] and r['doi'].startswith('10.'):
        lines.append(f"  doi = {{{r['doi']}}},")
    elif u:
        lines.append(f"  url = {{{u}}},")
    lines.append("}")
    return '\n'.join(lines)

def main(xlsx):
    wb = load_workbook(xlsx, data_only=True)
    ws = wb['Publicaciones']
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        n, anio, tipo, ref, doi = row[0], row[1], row[2], row[3], row[4]
        if isinstance(n, (int, float)) and isinstance(anio, (int, float)) and ref:
            rows.append({'n': int(n), 'anio': int(anio), 'tipo': str(tipo).strip(),
                         'ref': str(ref).strip(), 'doi': str(doi).strip() if doi else None})
    n_j = sum(1 for r in rows if r['tipo'] == 'Revista')
    n_c = len(rows) - n_j

    out = [HEAD.format(n_t=len(rows), n_j=n_j, n_c=n_c)]
    for tipo, h2 in [('Revista', 'Revistas'), ('Conferencia', 'Conferencias internacionales')]:
        cls = '' if tipo == 'Revista' else ' class="alt"'
        pad = ' style="padding-top:40px"' if tipo == 'Revista' else ''
        out.append(f'<section{cls}{pad}><div class="wrap"><div class="head"><h2>{h2}</h2></div>')
        year = None
        for r in rows:
            if r['tipo'] != tipo:
                continue
            if r['anio'] != year:
                year = r['anio']
                out.append(f'<div class="pubyear">{year}</div>')
            body = render_ref(r['ref'])
            u = doi_url(r['doi'])
            link = f'<a href="{html.escape(u)}" target="_blank" rel="noopener">DOI</a>' if u else ''
            out.append(f'<div class="pub"><div class="thumb">{thumb_for(r["ref"], tipo)}</div>'
                       f'<div><p>{body}</p><div class="links">{link}<span>{r["anio"]}</span></div></div></div>')
        out.append('</div></section>')
    out.append(FOOT)
    open('publicaciones.html', 'w').write('\n'.join(out))

    used = set()
    bib = ['% Publicaciones de Matías Díaz — generado desde Publicaciones_ListadoIEEE_v4.xlsx (agosto 2026).\n']
    for r in rows:
        bib.append(bib_entry(r, used))
    open('publications.bib', 'w').write('\n\n'.join(bib) + '\n')
    print(f'{len(rows)} publicaciones ({n_j} revistas, {n_c} conferencias)')

if __name__ == '__main__':
    main(sys.argv[1])
