#!/usr/bin/env python3
"""Genera publicaciones.html y publications.bib desde los antecedentes de la jerarquización.

Uso: python3 tools/gen_publicaciones.py <Listado_Publicaciones.xlsx> <Indice_Pubs_Nuevas.xlsx> <Formulario_vfinal.docx>

El formulario aporta el registro completo de carrera (tabla de publicaciones, 2013-2026);
el listado y el índice aportan el período 2021-2026 con DOI y clasificación WoS/conferencia.
Cada entrada se presenta con estructura de referencia IEEE: autores (nombre propio en
negrita), título, revista o conferencia, cita y enlace DOI cuando existe.
"""
import sys, html, re, unicodedata, json, os
from openpyxl import load_workbook
import docx

HEAD = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Publicaciones · Matías Díaz</title>
<meta name="description" content="Publicaciones de Matías Díaz (DIE USACH): artículos en revistas WoS-JCR y conferencias internacionales, con enlace DOI.">
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Instrument+Serif:ital@0;1&family=Archivo:wght@400;500;600;700&display=swap">
<link rel="stylesheet" href="assets/site.css">
</head>
<body>
<nav><div class="wrap">
  <a class="logo" href="index.html">Matías <b>Díaz</b></a>
  <a href="investigacion.html">Investigación</a>
  <a class="on" href="publicaciones.html">Publicaciones</a>
  <a href="videos.html">Videos</a>
  <a href="prensa.html">Prensa</a>
  <a href="docencia.html">Docencia</a>
  <a href="equipo.html">Equipo</a>
  <a href="postula.html">Postula</a>
  <a href="cv.html">CV</a>
</div></nav>
<header class="pagehead wrap">
  <h1>Publicaciones</h1>
  <p>{n_t} publicaciones científicas: {n_j} artículos en revistas WoS-JCR y {n_c} en conferencias internacionales. Perfil completo en <a href="https://scholar.google.com/citations?user=-43YaJIAAAAJ" target="_blank" rel="noopener">Google Scholar</a> · <a href="publications.bib">BibTeX</a>.</p>
</header>
"""

FOOT = """
<footer><div class="wrap">
  <span>© 2026 Matías Díaz · DIE USACH</span>
  <span>Generado desde el índice consolidado de publicaciones · <a href="publications.bib">publications.bib</a></span>
</div></footer>
<script src="assets/site.js"></script>
</body>
</html>
"""

CONF_KEYS = ("CONFERENCE", "PROCEEDINGS", "CONGRESS", "SYMPOSIUM", "WORKSHOP", "IECON",
             "ICIT", "EVER", "CHILECON", "ICA-ACCA", "ICAACCA", "COBEP", "SPEC", "ARGENCON",
             "PEDG", "ISIE", "EPE", "ECCE", "APEC", "IEEE CONF", "SEMINAR", "MEETING", "EXHIBITION")
JOUR_KEYS = ("TRANSACTIONS", "JOURNAL", "ENERGIES", "IET ", "ACCESS", "MAGAZINE",
             "PROCESSES", "MATHEMATICS", "RESULTS IN", "APPLIED", "PHOTOENERGY")

def esc(x):
    return html.escape(str(x).strip()) if x else ""

def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", "", s)

def bold_me(authors):
    """Negrita para el nombre propio en la lista de autores (ya escapada)."""
    return re.sub(r"((?:M\.?\s*)?D[ií]az(?:-Diaz)?,?\s*M\.?|M\.?\s+D[ií]az)", r"<b>\1</b>", authors)

ACRON = ("IEEE","IET","AC","DC","EV","V2G","MMC","M3C","MPC","PWM","SVM","HVDC","VSC","LVRT",
         "PV","XXIII","ICA","ACCA","CHILECON","IECON","ICIT","EVER","SPEC","COBEP","ARGENCON",
         "PEDG","ISIE","EPE","ECCE","APEC","USA","UK","II","III","IV")
def fix_case(v):
    if not v or sum(c.isupper() for c in v if c.isalpha()) < 0.7 * max(1, sum(c.isalpha() for c in v)):
        return v
    words = v.title().split()
    out = []
    for w in words:
        wu = re.sub(r"[^A-Za-z0-9]", "", w).upper()
        out.append(w.upper() if wu in ACRON else w)
    return " ".join(out)

def ieee_line(authors, title, venue, cita, year, url):
    parts = []
    if authors:
        parts.append(bold_me(esc(authors.rstrip(".,; "))) + ",")
    parts.append(f"“{esc(title.rstrip('. '))},”")
    if venue:
        parts.append(f"<em>{esc(fix_case(venue.strip(' .,;')))}</em>,")
    if cita:
        parts.append(f"{esc(cita)},")
    parts.append(f"{year}.")
    doi = f'{link(url, "DOI")}' if url else ""
    return " ".join(parts), doi

def link(url, label):
    if url and str(url).startswith("http"):
        return f'<a href="{esc(url)}" target="_blank" rel="noopener">{label}</a>'
    return ""

def bibkey(title, year):
    t = unicodedata.normalize("NFKD", title).encode("ascii", "ignore").decode()
    words = [w.lower() for w in re.findall(r"[A-Za-z]{4,}", t)][:3]
    return f"diaz{year}" + "".join(w[:4] for w in words)


def strip_urls(x):
    return re.sub(r"\s*🔗?\s*https?://\S+", "", x or "").strip(" .,;\n")

def find_norm(hay, needle, nchars=30):
    """Posición en `hay` donde parte `needle`, comparando normalizado. -1 si no está."""
    nh, mapping = [], []
    for i, ch in enumerate(hay):
        c = unicodedata.normalize("NFKD", ch).encode("ascii", "ignore").decode().lower()
        c = re.sub(r"[^a-z0-9]", "", c)
        if c:
            nh.append(c[0]); mapping.append(i)
    nn = norm(needle)[:nchars]
    if not nn:
        return -1
    pos = "".join(nh).find(nn)
    return mapping[pos] if pos >= 0 else -1

def load_authors_index(indice):
    """Mapa título normalizado -> lista de autores, desde el índice de publicaciones nuevas."""
    wb = load_workbook(indice, data_only=True)
    ws = wb.active
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        n, year, tipo, title, url, ref = (list(r) + [None] * 8)[:6]
        if not (title and ref):
            continue
        ref = re.sub(r"\s+", " ", strip_urls(str(ref)))
        key = norm(title)
        pos = find_norm(ref, str(title))
        if pos > 5:
            authors = ref[:pos].strip(" .,")
        elif pos == -1:
            authors = ref  # ref sin título: solo autores
        else:
            authors = ref[pos + len(str(title)):].strip(" .,")  # formato título-primero
        out[key] = authors
    return out

def lookup_authors(ix, title):
    key = norm(title)
    if key in ix:
        return ix[key]
    for k, v in ix.items():
        if k[:25] and (k.startswith(key[:25]) or key.startswith(k[:25])):
            return v
    return ""

def prev_pubs(formulario):
    """Registro de carrera con fecha hasta 2020: (año, autores, título, venue)."""
    doc = docx.Document(formulario)
    tab = None
    for t in doc.tables:
        head = " ".join(c.text for c in t.rows[1].cells) if len(t.rows) > 1 else ""
        if "NOMBRE DE LA PUBLICACI" in head and len(t.rows) > 100:
            tab = t
            break
    out = []
    for r in tab.rows[2:]:
        cells = [c.text.strip() for c in r.cells]
        if not cells[0]:
            continue
        m = re.search(r"(19|20)\d{2}", cells[5])
        yr = int(m.group(0)) if m else None
        if yr is None or yr > 2020:
            continue
        title_raw = re.sub(r"\s+", " ", cells[0]).strip()
        ref = re.sub(r"\s+", " ", strip_urls(cells[4]))
        authors, title, venue = "", title_raw.title(), ""
        if ref:
            pos = find_norm(ref, title_raw)
            if pos > 5:
                authors = ref[:pos].strip(" .,")
                title = ref[pos:pos + len(title_raw)].strip(" .")
                venue = ref[pos + len(title_raw):].strip(" .,")
            elif pos == 0:
                title = ref[:len(title_raw)].strip(" .")
                venue = ref[len(title_raw):].strip(" .,")
            else:
                authors = ref
        out.append((yr, authors, title, venue))
    return out

def classify(venue, title=""):
    v = (venue + " " + title).upper()
    if any(k in v for k in CONF_KEYS):
        return "conf"
    if any(k in v for k in JOUR_KEYS):
        return "journal"
    return "conf"

CODE_NAMES = {
    "ICAACCA": "IEEE International Conference on Automation / Congress of the Chilean Association of Automatic Control (ICA-ACCA)",
    "ICA-ACCA": "IEEE International Conference on Automation / Congress of the Chilean Association of Automatic Control (ICA-ACCA)",
    "CHILECON": "IEEE CHILEAN Conference on Electrical, Electronics Engineering, Information and Communication Technologies (CHILECON)",
    "EVER": "International Conference on Ecological Vehicles and Renewable Energies (EVER)",
    "ARGENCON": "IEEE Biennial Congress of Argentina (ARGENCON)",
    "SMART": "International Conference on Sustainable Mobility Applications, Renewables and Technology (SMART)",
    "INGELECTRA": "Congreso INGELECTRA",
    "IECON": "Annual Conference of the IEEE Industrial Electronics Society (IECON)",
    "ICIT": "IEEE International Conference on Industrial Technology (ICIT)",
    "SPEC": "IEEE Southern Power Electronics Conference (SPEC)",
}
_vjson = os.path.join(os.path.dirname(os.path.abspath(__file__)), "venues_conferencias.json")
VENUES = json.load(open(_vjson)) if os.path.exists(_vjson) else {}

def conf_venue(title, fuente, url):
    k = norm(title)[:48]
    if k in VENUES:
        return VENUES[k]["venue"], (VENUES[k]["doi"] or None)
    m = re.search(r"10\.1109/([A-Za-z-]+)", str(url or ""))
    if m and m.group(1).upper() in CODE_NAMES:
        return CODE_NAMES[m.group(1).upper()], None
    f = str(fuente or "").strip()
    return ("IEEE Conference" if f in ("", "IEEE Conf.", "IEEE Conf") else f), None

def sanitize(authors, venue):
    authors = re.sub(r"https?://\S+", "", authors or "").strip(" .,;")
    venue = re.sub(r"https?://\S+", "", venue or "").strip(" .,;")
    venue = re.sub(r"\s+", " ", venue)
    authors = re.sub(r"\s+", " ", authors)
    up = venue.upper()
    is_venue = any(k in up for k in CONF_KEYS + JOUR_KEYS + ("IEEE", "INTERNATIONAL", "REVISTA", "SPRINGER", "MDPI", "ELSEVIER"))
    # venue que en realidad es lista de autores (refs con formato titulo-primero)
    if venue and not is_venue and len(re.findall(r",\s*[A-Z]", venue)) >= 2:
        if len(authors) < 6:
            authors, venue = venue, ""
    if len(authors) < 6:
        authors = ""
    return authors, venue

# Correcciones puntuales verificadas en Crossref (título -> autores);
# y DOI incorrecto detectado: el de la review de MMC apunta a otro paper.
AUTHOR_OVERRIDES = {
    "conditionmonitoringofsubmodulecapacitors": "Saravanakumar R., Sivakumar N., Kirthika Devi V.S., Shanthini C., Jena D., Ibaceta E., Diaz M., Rodriguez J.",
}
DOI_DROP = {"reviewofmodularmultilevelconvertersappliedtohigh": "https://doi.org/10.1109/CHILECON47746.2019.8988098"}

def render_entry(out, authors, title, venue, cita, year, url):
    k = norm(title)[:48]
    for ok, ov in AUTHOR_OVERRIDES.items():
        if k.startswith(ok):
            authors = ov
    for dk, du in DOI_DROP.items():
        if k.startswith(dk) and url == du:
            url = None
    authors, venue = sanitize(authors, venue)
    body, doi = ieee_line(authors, title, venue, cita, year, url)
    links = f'<div class="links">{doi}<span>{year}</span></div>' if doi else f'<div class="links"><span>{year}</span></div>'
    thumb = esc(str(venue).split()[0][:10]) if venue else str(year)
    out.append(f'<div class="pub"><div class="thumb">{thumb}</div><div><p>{body}</p>{links}</div></div>')

def main(listado, indice, formulario):
    wb = load_workbook(listado, data_only=True)

    def _yr(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    authors_ix = load_authors_index(indice)
    journals, confs = [], []

    for r in wb["Revistas WoS-JCR"].iter_rows(min_row=2, values_only=True):
        if not (_yr(r[1]) and r[2]):
            continue
        _, y, title, journal, cita, _, url = (list(r) + [None] * 7)[:7]
        journal = journal if journal and "verificar" not in str(journal) else ""
        journals.append((int(y), lookup_authors(authors_ix, title), str(title), str(journal), cita, url))

    for r in wb["Conferencias-Proceedings"].iter_rows(min_row=2, values_only=True):
        if not (_yr(r[1]) and r[2]):
            continue
        _, y, title, fuente, url = (list(r) + [None] * 5)[:5]
        venue, doi2 = conf_venue(title, fuente, url)
        confs.append((int(y), lookup_authors(authors_ix, title), str(title), venue, None, url or doi2))

    for yr, authors, title, venue in prev_pubs(formulario):
        entry = (yr, authors, title, venue, None, None)
        (journals if classify(venue, title) == "journal" else confs).append(entry)

    journals.sort(key=lambda e: -e[0])
    confs.sort(key=lambda e: -e[0])

    out = [HEAD.format(n_j=len(journals), n_c=len(confs), n_t=len(journals) + len(confs))]

    out.append('<section style="padding-top:40px"><div class="wrap"><div class="head"><h2>Revistas WoS-JCR</h2></div>')
    for year in sorted({e[0] for e in journals}, reverse=True):
        out.append(f'<div class="pubyear">{year}</div>')
        for e in journals:
            if e[0] == year:
                render_entry(out, e[1], e[2], e[3], e[4], e[0], e[5])
    out.append("</div></section>")

    out.append('<section class="alt"><div class="wrap"><div class="head"><h2>Conferencias internacionales</h2></div>')
    for year in sorted({e[0] for e in confs}, reverse=True):
        out.append(f'<div class="pubyear">{year}</div>')
        for e in confs:
            if e[0] == year:
                render_entry(out, e[1], e[2], e[3], e[4], e[0], e[5])
    out.append("</div></section>")
    out.append(FOOT)

    with open("publicaciones.html", "w", encoding="utf-8") as f:
        f.write("\n".join(out))

    # BibTeX desde el índice de publicaciones nuevas (referencia completa + DOI)
    wb2 = load_workbook(indice, data_only=True)
    ws = wb2.active
    bibs, seen = [], set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        n, year, tipo, title, url, ref = (list(r) + [None] * 8)[:6]
        if not (year and title):
            continue
        key = bibkey(str(title), year)
        while key in seen:
            key += "x"
        seen.add(key)
        entry = "article" if tipo and ("WOS" in str(tipo).upper() or "revista" in str(tipo).lower()) else "inproceedings"
        doi = ""
        if url and "doi.org/" in str(url):
            doi = str(url).split("doi.org/")[-1].strip()
        bibs.append(
            f"@{entry}{{{key},\n"
            f"  title = {{{str(title).strip()}}},\n"
            f"  year = {{{year}}},\n"
            + (f"  doi = {{{doi}}},\n" if doi else "")
            + (f"  note = {{{str(ref).strip()[:300]}}},\n" if ref else "")
            + "}\n")
    with open("publications.bib", "w", encoding="utf-8") as f:
        f.write("% Publicaciones de Matías Díaz.\n"
                "% Generado desde 00_INDICE_Publicaciones_Nuevas_2021-2026.xlsx.\n\n" + "\n".join(bibs))
    print(f"OK: {len(journals)} revistas, {len(confs)} conferencias, total {len(journals)+len(confs)}, {len(bibs)} BibTeX")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2], sys.argv[3])
